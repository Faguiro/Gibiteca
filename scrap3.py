import openpyxl
import requests
from bs4 import BeautifulSoup
import re

def remove_special_characters(text):
    return re.sub(r'[<>:"/\\|?*]', '', text)

# Carregar o arquivo Excel original
workbook = openpyxl.load_workbook("comics_data_filtrado.xlsx")
sheet = workbook.active

# Criar uma nova planilha para armazenar os dados dos links de download
links_workbook = openpyxl.Workbook()
links_sheet = links_workbook.active

# Criar cabeçalho para a planilha de links
links_sheet.append(["Título do Comic", "Link de Download"])

# Iterar pelas linhas do arquivo original e acessar os links
for row in sheet.iter_rows(min_row=2, values_only=True):
    _, _, _, title, link = row  # Desempacotar os valores da linha
    
    # Acessar o link
    response = requests.get(link)
    if response.status_code == 200:
        try: 
            soup = BeautifulSoup(response.text, "html.parser")
            print(f"Coletando dados da página {link}")
            
            # Procurar por tags <ul> dentro da seção 'post-contents'
            post_contents = soup.find("section", class_="post-contents")
            if post_contents:
                ul_tags = post_contents.find_all("ul")
                for ul_tag in ul_tags:
                    li_tags = ul_tag.find_all("li")
                    for li_tag in li_tags:
                        # Extrair informações das tags <li>
                        li_text = remove_special_characters(li_tag.get_text())
                        
                        # Adicionar as informações à planilha de links
                        links_sheet.append([title, li_text])
        
            
        except Exception as e:
            print(f"Erro ao coletar dados da página {link}")
            print(e)

        links_workbook.save("comic_links.xlsx")
# Salvar a planilha de links
links_workbook.save("comic_links.xlsx")
print("Links de download salvos em 'comic_links.xlsx'")
