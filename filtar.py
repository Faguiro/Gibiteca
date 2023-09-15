import openpyxl

# Carregar o arquivo Excel
workbook = openpyxl.load_workbook("comics_data.xlsx")
sheet = workbook.active

# Criar uma nova planilha para armazenar os dados filtrados
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Copiar o cabeçalho para a nova planilha
new_sheet.append(["ID", "Capa", "Editora", "Título", "Link"])

# Iterar pelas linhas do arquivo original e adicionar apenas as que têm título não vazio
for row in sheet.iter_rows(min_row=2, values_only=True):
    _, _, _, title, _ = row  # Desempacotar os valores da linha
    if title:
        new_sheet.append(row)

# Salvar a nova planilha em um novo arquivo
new_workbook.save("comics_data_filtrado.xlsx")
print("Dados filtrados salvos em 'comics_data_filtrado.xlsx'")
