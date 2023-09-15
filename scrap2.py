import openpyxl
import requests
from bs4 import BeautifulSoup
import re

import sqlite3


# Conectar ao banco de dados SQLite ou criar um novo se não existir
conn = sqlite3.connect("comics_data.db")
cursor = conn.cursor()

# Criar a tabela se ela não existir
cursor.execute('''CREATE TABLE IF NOT EXISTS comics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    capa TEXT,
                    editora TEXT,
                    titulo TEXT,
                    link TEXT,
                    descricao TEXT,
                    links_adicionais TEXT
                )''')

def remove_special_characters(text):
    return re.sub(r'[<>:"/\\|?*]', '', text)

# Carregar o arquivo Excel original
workbook = openpyxl.load_workbook("comics_data_filtrado.xlsx")
sheet = workbook.active

# Criar uma nova planilha para armazenar os dados adicionais
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Copiar o cabeçalho original para a nova planilha e adicionar novas colunas
new_sheet.append(["ID", "Capa", "Editora", "Título", "Link", "Descrição", "Links Adicionais"])

for row in sheet.iter_rows(min_row=2, values_only=True):
    _, capa, editora, titulo, link = row  # Desempacotar os valores da linha
    # Acessar o link
    try:
        response = requests.get(link)
        response.raise_for_status()  # Verificar se há erros na resposta
    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar o link {link}: {e}")
        continue  # Continuar com a próxima iteração


    if response.status_code == 200:
        try: 
            soup = BeautifulSoup(response.text, "html.parser")
            #print(f"Coletando dados da página {link}")
            
            # Extrair o texto da primeira tag <p> dentro da section 'post-contents'
            post_contents = soup.find("section", class_="post-contents")
            if post_contents:
                first_p = post_contents.find("p")
                if first_p:
                    descricao = remove_special_characters(first_p.get_text())
                else:
                    descricao = ""
            else:
                descricao = ""
            
            # Encontrar todos os links da tag <ul> dentro da section 'post-contents'
            first_ul = post_contents.find("ul")     
            if first_ul:       
                colunas_adicionais = []
                for tag_a in first_ul.find_all("a"):
                    coluna_nome = tag_a.text.strip()
                    coluna_link = tag_a.get("href", "")
                    colunas_adicionais.append(f"{coluna_nome}: {coluna_link}")
                    #print(f"Coluna adicional: {coluna_nome} - {coluna_link}")
                links_adicionais = "\n".join(colunas_adicionais)
            else:
                links_adicionais = ""
            
            # Encontrar todas as divs com a classe 'aio-pulse'
            aio_pulse_divs = soup.find_all("div", class_="aio-pulse")
            colunas_adicionais = []
            for aio_pulse in aio_pulse_divs:
                a_tag = aio_pulse.find("a")
                if a_tag:
                    coluna_nome = a_tag.get("title", "")
                    coluna_link = a_tag.get("href", "")
                    colunas_adicionais.append(f"{coluna_nome}: {coluna_link}")
                    #print(f"Coluna adicional: {coluna_nome} - {coluna_link}")
                links_adicionais = "\n".join(colunas_adicionais)
            
            # Adicionar as informações à nova planilha
            new_row = list(row) + [descricao, links_adicionais]
            new_sheet.append(new_row)

            # Inserir os dados no banco de dados SQLite
            cursor.execute('''INSERT INTO comics (capa, editora, titulo, link, descricao, links_adicionais)
                              VALUES (?, ?, ?, ?, ?, ?)''', (capa, editora, titulo, link, descricao, links_adicionais))
            conn.commit()
            print("Registro n. {} inserido com sucesso".format(cursor.lastrowid))
        except Exception as e:
            print(f"Erro ao coletar dados da página {link}")
            print(e)
    

# Salvar a nova planilha com as informações adicionais
conn.close()
print('Total de registros inseridos no banco de dados:', cursor.lastrowid)
print("Dados com descrição e links adicionais salvos em 'comics_data_com_descricao.xlsx'")
